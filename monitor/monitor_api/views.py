from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.utils.dateparse import parse_datetime
from django.utils import timezone
from datetime import timedelta, datetime
from metrics.models import Metric, Host
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from io import BytesIO

def home(request):
    return HttpResponse("<h1>API de Monitoramento — OK</h1><p>Use /api/</p>")

def dashboard(request):
    return render(request, "dashboard.html")

def generate_report(request):
    """
    ✅ CORREÇÃO: Gera relatórios em PDF ou XLSX com timezone correto
    CRÍTICO: NOW sempre UTC-aware para queries no banco
    """
    host_id = request.GET.get("host")
    range_param = request.GET.get("range", "24h")
    format_param = request.GET.get("format", "xlsx")
    
    start_custom = request.GET.get("start_date")
    end_custom = request.GET.get("end_date")

    if not host_id:
        return HttpResponse("Host ID é obrigatório", status=400)

    try:
        host = Host.objects.get(id=host_id)
    except Host.DoesNotExist:
        return HttpResponse("Host não encontrado", status=404)

    # ✅ NOW sempre UTC-aware para queries
    now = timezone.now()
    
    print(f"\n{'='*80}")
    print(f"[GENERATE_REPORT] Range: {range_param}")
    print(f"[GENERATE_REPORT] NOW (UTC): {now}")
    print(f"[GENERATE_REPORT] NOW (Local): {timezone.localtime(now)}")
    
    if range_param == 'custom' and start_custom and end_custom:
        try:
            start_time = parse_datetime(start_custom)
            end_time = parse_datetime(end_custom)
            
            # ✅ Garantir que são timezone-aware
            if start_time and timezone.is_naive(start_time):
                start_time = timezone.make_aware(start_time)
            if end_time and timezone.is_naive(end_time):
                end_time = timezone.make_aware(end_time)
                
            print(f"[GENERATE_REPORT] CUSTOM (UTC): {start_time} até {end_time}")
        except ValueError:
            start_time = now - timedelta(hours=24)
            end_time = now
            print(f"[GENERATE_REPORT] Erro parsing custom, usando 24h padrão")
    else:
        # ✅ Presets com NOW em UTC
        end_time = now
        
        if range_param == '1h':
            start_time = now - timedelta(hours=1)
        elif range_param == '6h':
            start_time = now - timedelta(hours=6)
        elif range_param == '7d':
            start_time = now - timedelta(days=7)
        else:  # 24h default
            start_time = now - timedelta(hours=24)
        
        print(f"[GENERATE_REPORT] START (UTC): {start_time}")
        print(f"[GENERATE_REPORT] END (UTC): {end_time}")
        print(f"[GENERATE_REPORT] Diferença: {(end_time - start_time).total_seconds() / 3600:.1f} horas")

    # ✅ Busca no banco com filtro correto (UTC-aware)
    metrics = Metric.objects.filter(
        host=host,
        timestamp__gte=start_time,
        timestamp__lte=end_time
    ).order_by('timestamp')

    count = metrics.count()
    print(f"[GENERATE_REPORT] Total encontrado: {count}")
    print(f"{'='*80}\n")

    # PREPARAÇÃO DOS DADOS (CONVERSÃO PARA LOCAL TIME apenas para exibição)
    cpu_data = []
    memory_data = []

    for m in metrics:
        # ✅ Converte de UTC (Banco) para Local (Brasil) APENAS para exibição
        local_ts = timezone.localtime(m.timestamp)
        
        if m.metric_type == 'cpu_percent':
            cpu_data.append((local_ts, m.value))
        elif m.metric_type == 'memory_percent':
            memory_data.append((local_ts, m.value))

    # Gera o arquivo
    if format_param == 'pdf':
        return generate_pdf_report(host, cpu_data, memory_data, range_param)
    else:
        return generate_xlsx_report(host, cpu_data, memory_data, range_param)


def generate_xlsx_report(host, cpu_data, memory_data, range_param):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório"

    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    center_align = Alignment(horizontal="center", vertical="center")
    border_style = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    bold_font = Font(bold=True)

    # Cabeçalho Principal
    ws.merge_cells('A1:D1')
    ws['A1'] = f"Relatório de Monitoramento - {host.hostname}"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align

    # Metadados
    now_local = timezone.localtime(timezone.now())
    ws['A3'] = f"Host: {host.hostname}"
    ws['A4'] = f"IP: {host.ip if host.ip else 'N/A'}"
    ws['A5'] = f"Intervalo: {range_param}"
    ws['A6'] = f"Gerado em: {now_local.strftime('%d/%m/%Y %H:%M:%S')}"

    # ==========================================
    # TABELA CPU
    # ==========================================
    current_row = 8
    
    # Título
    ws.merge_cells(f'A{current_row}:B{current_row}')
    cell_title = ws[f'A{current_row}']
    cell_title.value = "DADOS DE CPU (%)"
    cell_title.font = Font(bold=True, color="FFFFFF")
    cell_title.fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
    cell_title.alignment = center_align
    ws[f'A{current_row}'].border = border_style
    ws[f'B{current_row}'].border = border_style
    
    current_row += 1
    
    # Cabeçalhos das Colunas
    headers = ["Data/Hora", "Valor (%)"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border_style

    current_row += 1
    
    # Dados CPU
    if not cpu_data:
        ws.merge_cells(f'A{current_row}:B{current_row}')
        ws.cell(row=current_row, column=1, value="Sem dados no período").alignment = center_align
        ws.cell(row=current_row, column=1).border = border_style
        ws.cell(row=current_row, column=2).border = border_style
        current_row += 1
    else:
        cpu_values = []
        for ts, val in cpu_data:
            cpu_values.append(val)
            # Data
            c1 = ws.cell(row=current_row, column=1, value=ts.replace(tzinfo=None))
            c1.number_format = 'dd/mm/yyyy hh:mm:ss'
            c1.border = border_style
            c1.alignment = center_align
            # Valor
            c2 = ws.cell(row=current_row, column=2, value=val)
            c2.border = border_style
            c2.alignment = center_align
            current_row += 1

        # --- ESTATÍSTICAS CPU ---
        # Mínimo
        ws.cell(row=current_row, column=1, value="Mínimo").font = bold_font
        ws.cell(row=current_row, column=1).border = border_style
        ws.cell(row=current_row, column=1).alignment = center_align
        
        ws.cell(row=current_row, column=2, value=min(cpu_values)).font = bold_font
        ws.cell(row=current_row, column=2).border = border_style
        ws.cell(row=current_row, column=2).alignment = center_align
        current_row += 1

        # Máximo
        ws.cell(row=current_row, column=1, value="Máximo").font = bold_font
        ws.cell(row=current_row, column=1).border = border_style
        ws.cell(row=current_row, column=1).alignment = center_align

        ws.cell(row=current_row, column=2, value=max(cpu_values)).font = bold_font
        ws.cell(row=current_row, column=2).border = border_style
        ws.cell(row=current_row, column=2).alignment = center_align
        current_row += 1

        # Média
        ws.cell(row=current_row, column=1, value="Média").font = bold_font
        ws.cell(row=current_row, column=1).border = border_style
        ws.cell(row=current_row, column=1).alignment = center_align

        avg = sum(cpu_values) / len(cpu_values)
        ws.cell(row=current_row, column=2, value=round(avg, 2)).font = bold_font
        ws.cell(row=current_row, column=2).border = border_style
        ws.cell(row=current_row, column=2).alignment = center_align
        current_row += 1

    # ==========================================
    # TABELA MEMÓRIA
    # ==========================================
    current_row += 2
    
    # Título
    ws.merge_cells(f'A{current_row}:B{current_row}')
    cell_title = ws[f'A{current_row}']
    cell_title.value = "DADOS DE MEMÓRIA RAM (%)"
    cell_title.font = Font(bold=True, color="FFFFFF")
    cell_title.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    cell_title.alignment = center_align
    ws[f'A{current_row}'].border = border_style
    ws[f'B{current_row}'].border = border_style

    current_row += 1
    
    # Cabeçalhos
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border_style

    current_row += 1
    
    # Dados Memória
    if not memory_data:
        ws.merge_cells(f'A{current_row}:B{current_row}')
        ws.cell(row=current_row, column=1, value="Sem dados no período").alignment = center_align
        ws.cell(row=current_row, column=1).border = border_style
        ws.cell(row=current_row, column=2).border = border_style
    else:
        mem_values = []
        for ts, val in memory_data:
            mem_values.append(val)
            # Data
            c1 = ws.cell(row=current_row, column=1, value=ts.replace(tzinfo=None))
            c1.number_format = 'dd/mm/yyyy hh:mm:ss'
            c1.border = border_style
            c1.alignment = center_align
            # Valor
            c2 = ws.cell(row=current_row, column=2, value=val)
            c2.border = border_style
            c2.alignment = center_align
            current_row += 1

        # --- ESTATÍSTICAS MEMÓRIA ---
        # Mínimo
        ws.cell(row=current_row, column=1, value="Mínimo").font = bold_font
        ws.cell(row=current_row, column=1).border = border_style
        ws.cell(row=current_row, column=1).alignment = center_align
        
        ws.cell(row=current_row, column=2, value=min(mem_values)).font = bold_font
        ws.cell(row=current_row, column=2).border = border_style
        ws.cell(row=current_row, column=2).alignment = center_align
        current_row += 1

        # Máximo
        ws.cell(row=current_row, column=1, value="Máximo").font = bold_font
        ws.cell(row=current_row, column=1).border = border_style
        ws.cell(row=current_row, column=1).alignment = center_align

        ws.cell(row=current_row, column=2, value=max(mem_values)).font = bold_font
        ws.cell(row=current_row, column=2).border = border_style
        ws.cell(row=current_row, column=2).alignment = center_align
        current_row += 1

        # Média
        ws.cell(row=current_row, column=1, value="Média").font = bold_font
        ws.cell(row=current_row, column=1).border = border_style
        ws.cell(row=current_row, column=1).alignment = center_align

        avg = sum(mem_values) / len(mem_values)
        ws.cell(row=current_row, column=2, value=round(avg, 2)).font = bold_font
        ws.cell(row=current_row, column=2).border = border_style
        ws.cell(row=current_row, column=2).alignment = center_align

    # Ajuste de largura
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"relatorio_{host.hostname}_{timezone.localtime(timezone.now()).strftime('%Y%m%d_%H%M')}.xlsx"
    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def generate_pdf_report(host, cpu_data, memory_data, range_param):
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=letter)
    story = []
    styles = getSampleStyleSheet()

    # Título
    story.append(Paragraph(f"Relatório de Monitoramento - {host.hostname}", styles['Title']))
    story.append(Spacer(1, 12))

    # Metadados
    now_local = timezone.localtime(timezone.now())
    meta_text = f"""
    <b>Host:</b> {host.hostname}<br/>
    <b>IP:</b> {host.ip or 'N/A'}<br/>
    <b>Período:</b> {range_param}<br/>
    <b>Gerado em:</b> {now_local.strftime('%d/%m/%Y %H:%M:%S')}
    """
    story.append(Paragraph(meta_text, styles['Normal']))
    story.append(Spacer(1, 20))

    # --- TABELA CPU ---
    story.append(Paragraph("<b>CPU (%)</b>", styles['Heading2']))
    
    if not cpu_data:
        story.append(Paragraph("Sem dados para o período selecionado.", styles['Normal']))
    else:
        data = [['Data/Hora', 'Valor (%)']]
        for ts, val in cpu_data[:500]: 
            data.append([ts.strftime('%d/%m/%Y %H:%M:%S'), f"{val:.2f}%"])

        vals = [v for _, v in cpu_data]
        data.append(['Mínimo', f"{min(vals):.2f}%"])
        data.append(['Máximo', f"{max(vals):.2f}%"])
        data.append(['Média', f"{sum(vals)/len(vals):.2f}%"])

        t = Table(data, colWidths=[3*inch, 1.5*inch])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, -3), (-1, -1), colors.lightgrey),
        ]))
        story.append(t)

    story.append(Spacer(1, 20))

    # --- TABELA MEMÓRIA ---
    story.append(Paragraph("<b>MEMÓRIA RAM (%)</b>", styles['Heading2']))

    if not memory_data:
        story.append(Paragraph("Sem dados para o período selecionado.", styles['Normal']))
    else:
        data = [['Data/Hora', 'Valor (%)']]
        for ts, val in memory_data[:500]:
            data.append([ts.strftime('%d/%m/%Y %H:%M:%S'), f"{val:.2f}%"])

        vals = [v for _, v in memory_data]
        data.append(['Mínimo', f"{min(vals):.2f}%"])
        data.append(['Máximo', f"{max(vals):.2f}%"])
        data.append(['Média', f"{sum(vals)/len(vals):.2f}%"])

        t = Table(data, colWidths=[3*inch, 1.5*inch])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, -3), (-1, -1), colors.lightgrey),
        ]))
        story.append(t)

    doc.build(story)
    output.seek(0)
    
    now_local = timezone.localtime(timezone.now())
    filename = f"relatorio_{host.hostname}_{now_local.strftime('%Y%m%d_%H%M')}.pdf"
    response = HttpResponse(output.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

def report(request):
    """
    ✅ CORREÇÃO: Retorna JSON para o Dashboard com filtro temporal CORRETO
    CRÍTICO: NOW sempre UTC-aware para queries
    """
    host = request.GET.get("host")
    range_param = request.GET.get("range", "24h")

    qs = Metric.objects.all().order_by('timestamp')
    
    if host:
        qs = qs.filter(host_id=host)

    # ✅ NOW sempre UTC-aware para queries no banco
    now = timezone.now()
    
    print(f"\n{'='*80}")
    print(f"[REPORT JSON] Range: {range_param}")
    print(f"[REPORT JSON] NOW (UTC): {now}")
    print(f"[REPORT JSON] NOW (Local): {timezone.localtime(now)}")
    
    # ✅ Calcula START baseado em NOW (UTC)
    if range_param == '1h':
        start_time = now - timedelta(hours=1)
    elif range_param == '6h':
        start_time = now - timedelta(hours=6)
    elif range_param == '7d':
        start_time = now - timedelta(days=7)
    else:
        start_time = now - timedelta(hours=24)

    print(f"[REPORT JSON] START (UTC): {start_time}")
    print(f"[REPORT JSON] END (UTC): {now}")
    print(f"[REPORT JSON] Diferença: {(now - start_time).total_seconds() / 3600:.1f} horas")

    # ✅ Filtro com NOW como referência
    qs = qs.filter(timestamp__gte=start_time, timestamp__lte=now)
    
    count = qs.count()
    print(f"[REPORT JSON] Total encontrado: {count}")
    print(f"{'='*80}\n")

    data = [
        {
            "timestamp": m.timestamp.isoformat(),
            "metric_type": m.metric_type,
            "value": m.value
        } for m in qs
    ]
    
    return JsonResponse({"report": data}, safe=False)
